import openpyxl
wb = openpyxl.load_workbook('/Users/antsa/Documents/Antsa/PROJECTS/CAN-QUEBEC/lnq-2026/public/data/HORAIRE_2026.xlsx', data_only=True)
print("Feuilles:", wb.sheetnames)
ws = wb.active
count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    c, t1, t2, co = row[11] if len(row)>11 else None, row[12] if len(row)>12 else None, row[13] if len(row)>13 else None, row[14] if len(row)>14 else None
    if c or t1 or t2 or co:
        print(f"C:{c} | T1:{t1} | T2:{t2} | CO:{co}")
        count += 1
        if count >= 20:
            break
